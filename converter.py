import os
from openpyxl import load_workbook
from json import load


def prefix_remover(path, prefix):
    for file in os.listdir(path):
        if file.startswith(prefix):
            new_name = file.replace(prefix, "")
            os.rename(os.path.join(path, file), os.path.join(path, new_name))


def converter(pics_path, excel_full_path, new_ext, prefix):
    print("please Wait")
    workbook = load_workbook(excel_full_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        for file_name in os.listdir(pics_path):
            prefix_remover(pics_path, prefix)
            file_raw_name = file_name.split(".")[0]
            if str(row[0]) == file_raw_name:
                new_name = (str(row[2]) + "." + new_ext)
                os.rename(os.path.join(pics_path, file_name),
                          os.path.join(pics_path, new_name))
    print("Done")


# pictures_path = "F:\\Programing\\projects\\behzisti\\pics"
# pictures_path = "D:\\Madakto\\person pic\\convertname" #**
# excel_path = "D:\\Madakto\\person pic\\convertname\\List.xlsx" #**
# excel_path = "F:\\Programing\\projects\\behzisti\\exel\\employee.xlsx"
# TODO: copy the pictures not just change it-- speak to user

with open('settings') as f:
    json_data = load(f)
    # print(json_data)

pictures_path = json_data["pictures path"]
excel_path =json_data["excel path"]
print(pictures_path ,excel_path)

# converter(pics_path=pictures_path,
#           excel_full_path=excel_path,
#           new_ext="png",
#           prefix="LF")
# jpg -> png


# def extension_changer(path, old_ext, new_ext):
#     for file in os.listdir(path):
#         if file.endswith(old_ext):
#             new_name = file.replace(old_ext, new_ext)
#             os.rename(os.path.join(path, file), os.path.join(path, new_name))